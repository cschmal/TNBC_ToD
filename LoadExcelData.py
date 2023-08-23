#!/usr/bin/env python

"""
This module loads experimental data as given in form of an excel sheet and
saves it as a pickle file
"""

__author__ = "Christoph Schmal"
__license__ = "GPL"
__maintainer__ = "Christoph Schmal"
__email__ = "christoph.schmal@hu-berlin.de"
__repository__ = "https://github.com/cschmal/TNBC_ToD"


from openpyxl import load_workbook
import pandas as pd
import pickle

def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

# ToDo: Last cell hard coded...any way to get that automatically?
file2load, last_cell = 'ToD_Data.xlsx', 15

wb = load_workbook(file2load)
ws = wb['Selected_Celllines']

cell = ws['H2']
print( cell )

# load headers from 1st line
headers = set([getMergedCellVal(ws, i) for i in ws['1']])
print("Headers", headers)
headers.remove(None)        # remove None values from the list
headers = list( headers )   # convert to list for item modification

compact_headers = headers.copy()
print("Headers 2", compact_headers)

for n, i in enumerate(compact_headers):
    compact_headers[n] = i.replace(' ', '_')

# initialize dicts from names in the set
for i in compact_headers:
    exec(i+r"={}")      # exec executes string as it were pythonic

#
# Load Classification
#

# Read in the cell ranges of the merged cell Classification
for rng in ws.merged_cells.ranges:
    if 'A1' in rng:
        break
else:
    rng = None

#
# Read in cell values and create pandas data frame
#

#print(rng, type(rng), rng.min_row, rng.max_row, rng.min_col, rng.max_col)
headers = [ws.cell(row=4, column=j).value for j in range(rng.min_col, rng.max_col+1)]
values  = [ [ws.cell(row=i, column=j).value for j in range(rng.min_col, rng.max_col+1)] for i in range(5, last_cell+1)]

#create pandas data frame from value list
Classification = pd.DataFrame(values, columns=headers, dtype=str)

#
# Load circadian/growth data results - 1st cells (classification) span 2 rows
#

for classification_name in ["Circadian Clock", "Growth Properties"]:
    #classification_name = "classification_name"
    # Find cell_ranges for header "classification_name"
    for rng in ws.merged_cells.ranges:
        print("CC+GP: ", rng, rng.start_cell.value )
        if rng.start_cell.value == classification_name:
            break
        else:
            pass

    # determine column ranges for property class "classification_name"
    col_min, col_max = rng.min_col, rng.max_col
    # load clock reporter / growth parameter names
    reporters = set([getMergedCellVal(ws, ws.cell(row=3, column=j)) for j in range(col_min, col_max+1)])
    if None in reporters:
        reporters.remove(None)
    reporters = list(reporters)
    #print( reporters )

    # Build separate pandas data frame with clock properies for each reporter
    # we store this data frame in the previosuly

    for c_reporter in reporters:
        #c_reporter = "Bmal1"
        # choose column indices whose values == i
        reporter_indices = [ j for j in range(col_min, col_max+1) if getMergedCellVal(ws, ws.cell(row=3, column=j)) == c_reporter ]
        #print( reporter_indices )

        # read in headers and values
        headers = [ws.cell(row=4, column=j).value for j in reporter_indices]
        values  = [ [ws.cell(row=i, column=j).value for j in reporter_indices] for i in range(5, last_cell+1)]
        #print( headers )
        #print( values )

        # create pandas data frame and store it in the corresponding dictionary
        reporter_values = pd.DataFrame(values, columns=headers, dtype=float)
        # insert classification info to the data frame
        for i in range(0, Classification.shape[1]):
            reporter_values.insert(i, Classification.keys()[i], Classification[Classification.keys()[i]])
        #print(reporter_values)
        globals()[classification_name.replace(" ", "_")][c_reporter] = reporter_values


for classification_name in ["Drug Sensitivity", "Time of Day Sensitivity"]:

    # Find cell_ranges for header "classification_name"
    for rng in ws.merged_cells.ranges:
        print("DS+TS: ", rng, rng.start_cell.value )
        if rng.start_cell.value == classification_name:
            break
        else:
            pass

    # determine column ranges for property class "classification_name"
    col_min, col_max = rng.min_col, rng.max_col
    #print(col_min, col_max)

    # load second level classification
    sec_level = set([getMergedCellVal(ws, ws.cell(row=2, column=j)) for j in range(col_min, col_max)])
    if None in reporters:
        sec_level.remove(None)
    sec_level = list(sec_level)

    # loop over second level name
    for sec_level_name in sec_level:

        globals()[classification_name.replace(" ", "_")][sec_level_name] = {}
        #print(sec_level_name)

        # Find cell_ranges for header "sec_level_name"
        for rng in ws.merged_cells.ranges:
            #print(rng, rng.start_cell.value )
            if rng.start_cell.value == sec_level_name:
                break
            else:
                pass
        # determine column ranges for property class "sec_level_name"
        col_min, col_max = rng.min_col, rng.max_col
        #print(col_min, col_max)

        # load third level classification
        third_level = set([getMergedCellVal(ws, ws.cell(row=3, column=j)) for j in range(col_min, col_max)])
        if None in reporters:
            third_level.remove(None)
        third_level = list(third_level)
        #print(third_level)

        # loop over third level entries
        for third_level_name in third_level:
            #third_level_name = third_level[0]
            #print(third_level_name)
            third_level_entry_indices = [j for j in range(col_min, col_max) if getMergedCellVal(ws, ws.cell(row=3, column=j)) == third_level_name]
            #third_level_col_min, third_level_col_max = min(third_level_entry_indices), max(third_level_entry_indices)
            headers = [ws.cell(row=4, column=j).value for j in third_level_entry_indices]
            values  = [ [ws.cell(row=i, column=j).value for j in third_level_entry_indices] for i in range(5, last_cell+1)]
            #print( headers )
            #print( values )
            c_values = pd.DataFrame(values, columns=headers, dtype=float)
            # insert classification info to the data frame
            for i in range(0, Classification.shape[1]):
                c_values.insert(i, Classification.keys()[i], Classification[Classification.keys()[i]])
            #print( c_values )
            globals()[classification_name.replace(" ", "_")][sec_level_name][third_level_name] = c_values


AllData2Pickle = Circadian_Clock, Growth_Properties, Drug_Sensitivity, Time_of_Day_Sensitivity

# Pickle data
with open('PickledData'+file2load.split('.')[0]+'.pickle', 'wb') as f:
    # Pickle the 'data' dictionary using the highest protocol available.
    pickle.dump(AllData2Pickle, f, pickle.HIGHEST_PROTOCOL)

# Check if pickling worked fine
with open('PickledData'+file2load.split('.')[0]+'.pickle', 'rb') as f:
    # The protocol version used is detected automatically, so we do not
    # have to specify it.
    data = pickle.load(f)
#print(data)
